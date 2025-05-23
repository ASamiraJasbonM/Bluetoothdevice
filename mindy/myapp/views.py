from django.shortcuts import render, HttpResponse
from django.shortcuts import render, redirect
from django import forms
# Create your views here.
from .models import PreguntaEncuesta, PreguntaEncuesta, RespuestaUsuario
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import timedelta
from datetime import datetime
from django.contrib import messages
from .forms import RegistroPorCedulaForm
from .forms import CrearCuentaForm
from django.contrib import messages
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import CrearCuentaForm
from .models import Cita, Usuarios, Usuariosparticipantes, Educativo, historialmed
from django.contrib.auth.decorators import login_required
import openpyxl
from django.http import HttpResponse
from io import BytesIO  # Asegúrate de importar esto
from openpyxl import Workbook
from django.http import HttpResponse
from .forms import HistorialmedForm, newpatient, EducativoForm
from .models import historialmed, Usuariosparticipantes
from django.contrib.auth import logout
from django.contrib.auth.hashers import check_password 
from django.contrib.auth.models import User
def politica_view(request):
    return render(request, 'politica.html')

def terminos_view(request):
    return render(request, 'terminos.html')


def crear_cuenta(request):
    if request.method == 'POST':
        form = CrearCuentaForm(request.POST)

        if form.is_valid():
            cedula = form.cleaned_data['cedula']

            # Get or create a user in Usuarios
            usuario, created = Usuarios.objects.get_or_create(cedula=cedula)

            # Create an auth user if not exists
            if not User.objects.filter(username=cedula).exists():
                user = User.objects.create_user(
                    username=cedula,
                    password=form.cleaned_data['contraseña'],
                    email=form.cleaned_data['correo1'],
                    first_name=form.cleaned_data['nombre'],
                    last_name=form.cleaned_data['apellidos']
                )
                user.save()

            # Update Usuarios fields
            usuario.nombre = form.cleaned_data['nombre']
            usuario.apellidos = form.cleaned_data['apellidos']
            usuario.correo = form.cleaned_data['correo1']
            usuario.fechanacimiento = form.cleaned_data['fechanacimiento']
            usuario.direccion = form.cleaned_data['direccion']
            usuario.contraseña = form.cleaned_data['contraseña']  # (Consider hashing this)
            usuario.spreguntas = form.cleaned_data['spreguntas']
            usuario.preguntas = form.cleaned_data['preguntas']
            usuario.save()

            messages.success(request, 'Cuenta creada exitosamente.')
            return redirect('home')
    else:
        form = CrearCuentaForm()

    return render(request, 'crear_cuenta.html', {'form': form})


def logout_view(request):
    logout(request)  # Eliminar la sesión del usuario
    return redirect('login')  # Redirigir al login después de cerrar sesión

'''def login_view(request): 
    if request.method == 'POST':
        correo = request.POST.get('correo')
        print(correo)
        password = request.POST.get('password')
        print(correo)
        usuarios = Usuarios.objects.filter(correo=correo)

        if usuarios.exists():
            usuario = usuarios.first()

            if usuario.contraseña == password:
                # Guardar el ID del usuario en la sesión
                print('coincide')
                request.session['usuario_id'] = usuario.cedula
                request.session['usuario_nombre'] = usuario.nombre
                request.session['usuario_apellidos'] = usuario.apellidos
                request.session['usuario_tipo'] = usuario.tipoa
                request.session['usuario_correo'] = usuario.correo
                request.session['usuario_direccion'] = usuario.direccion
                request.session['usuario_fechanacimiento'] = usuario.fechanacimiento

                # Verificar que la sesión se ha creado
                print("Sesión creada con éxito:", request.session.get('usuario_id'))

                # Redirección según el tipo de usuario
                if usuario.tipoa == "Participante":
                    print('participante')
                    return redirect('menu')  # O la vista correspondiente
                elif usuario.tipoa == "Administrador":
                    print('administrador')
                    return redirect('panel_admin')
                elif usuario.tipoa == "Profesional":
                    print('Participante')
                    return redirect('menu_profesional')
                else:
                    return redirect('menu')  # Redirigir a una vista genérica

            else:
                messages.error(request, 'Contraseña incorrecta.')
        else:
            messages.error(request, 'Usuario no encontrado.')

    return render(request, 'inicio_sesion.html')'''

@csrf_exempt
def login_view(request):
    if request.method == 'POST':
        print(request.POST)
        correo = request.POST.get('correo').lower()
        password = request.POST.get('password')

        usuario = Usuarios.objects.filter(correo__iexact=correo).first()
        
        if usuario:
            if password == usuario.contraseña:
                request.session['usuario_id'] = usuario.cedula
                request.session['usuario_nombre'] = usuario.nombre
                request.session['usuario_apellidos'] = usuario.apellidos
                request.session['usuario_tipo'] = usuario.tipoa
                request.session['usuario_correo'] = usuario.correo
                request.session['usuario_direccion'] = usuario.direccion
                request.session['usuario_fechanacimiento'] = str(usuario.fechanacimiento)

                if usuario.tipoa == "Participante":
                    return redirect('menu')
                elif usuario.tipoa == "Administrador":
                    return redirect('panel_admin')
                elif usuario.tipoa == "Profesional":
                    return redirect('menu_profesional')
                else:
                    return redirect('menu')
            else:
                messages.error(request, 'Contraseña incorrecta.')
        else:
            messages.error(request, 'Usuario no encontrado.')

    return render(request, 'inicio_sesion.html')


from .forms import CambiarContraseñaForm
def cambiar_contraseña(request):
    if 'usuario_id' not in request.session:
        return redirect('login')  # Redirige si no ha iniciado sesión

    try:
        usuario = Usuarios.objects.get(cedula=request.session['usuario_id'])
    except Usuarios.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('login')

    if request.method == 'POST':
        form = CambiarContraseñaForm(request.POST)
        if form.is_valid():
            actual = form.cleaned_data['contraseña_actual']
            nueva = form.cleaned_data['nueva_contraseña']
            confirmar = form.cleaned_data['confirmar_contraseña']

            # ⚠️ Verifica si la contraseña actual está en texto plano y la convierte a hash
            if not usuario.contraseña.startswith('pbkdf2_'):
                usuario.contraseña = make_password(usuario.contraseña)
                usuario.save()

            # ✅ Verificación segura de contraseña encriptada
            if not check_password(actual, usuario.contraseña):
                messages.error(request, 'La contraseña actual no es correcta.')
            elif nueva != confirmar:
                messages.error(request, 'Las nuevas contraseñas no coinciden.')
            else:
                usuario.contraseña = make_password(nueva)
                usuario.save()
                messages.success(request, 'Contraseña cambiada exitosamente.')
                return redirect('menu')
    else:
        form = CambiarContraseñaForm()

    return render(request, 'cambiocontraseña.html', {'form': form})

def ultima_encuesta(request):
    ultima_num_encuesta = PreguntaEncuesta.objects.latest('numero_encuesta').numero_encuesta
    preguntas = PreguntaEncuesta.objects.filter(numero_encuesta=ultima_num_encuesta).order_by('numero_pregunta')
    
    return render(request, 'ultima_encuesta.html', {
        'preguntas': preguntas,
        'numero_encuesta': ultima_num_encuesta,
    })



def citas_paciente(request):
    # Obtener la cédula del usuario actual desde la sesión (supongamos que ya está autenticado)
    cedula_usuario = request.session.get('usuario_id')

    # Obtener las citas del paciente
    citas = Cita.objects.filter(cedula=cedula_usuario)

    # Dividir las citas en asistidas, inasistidas y programadas
    citas_asistidas = []
    citas_inasistidas = []
    citas_programadas = []

    for cita in citas:
        # Calcular el tiempo restante para la cita
        fecha_cita = timezone.make_aware(datetime.combine(cita.dia, cita.hora))
        tiempo_restante = fecha_cita - timezone.now()

        # Clasificar las citas
        if tiempo_restante.total_seconds() < 0:
            # Si la cita ya pasó, la clasificamos como inasistida
            citas_inasistidas.append({
                'cita': cita,
                'tiempo_restante': tiempo_restante
            })
        elif tiempo_restante.total_seconds() > 0:
            # Si la cita está programada para el futuro
            citas_programadas.append({
                'cita': cita,
                'tiempo_restante': tiempo_restante
            })

        # Asistidas (si la cita está ya realizada, suponemos que la asistió)
        if cita.dia <= timezone.now().date() and cita.hora <= timezone.now().time():
            citas_asistidas.append({
                'cita': cita,
                'tiempo_restante': tiempo_restante
            })

    # Obtener el usuario participante y sus datos
    try:
        usuario_participante = Usuariosparticipantes.objects.get(cedula=cedula_usuario)
    except Usuariosparticipantes.DoesNotExist:
        usuario_participante = None

    return render(request, 'citas_paciente.html', {
        'citas_asistidas': citas_asistidas,
        'citas_inasistidas': citas_inasistidas,
        'citas_programadas': citas_programadas,
        'usuario_participante': usuario_participante
    })

def citas_profesional(request):
    # Obtener el nombre del profesional desde la sesión (supuesto que ya está autenticado)
    nombre_profesional = request.session.get('usuario_nombre')

    # Buscar las citas programadas para el profesional
    citas = Cita.objects.filter(profesional=nombre_profesional)

    citas_programadas = []

    for cita in citas:
        # Obtener la información del paciente
        paciente = Usuarios.objects.get(cedula=cita.cedula)

        # Obtener la edad del paciente
        fecha_nacimiento = paciente.fechanacimiento
        edad = calcular_edad(fecha_nacimiento)

        # Calcular el tiempo restante para la cita
        fecha_cita = timezone.make_aware(datetime.combine(cita.dia, cita.hora))
        tiempo_restante = fecha_cita - timezone.now()

        # Clasificar las citas que ya han pasado
        if tiempo_restante.total_seconds() < 0:
            citas_programadas.append({
                'cita': cita,
                'paciente_nombre': paciente.nombre,
                'paciente_edad': edad,
                'paciente_cedula': paciente.cedula,
                'tiempo_restante': tiempo_restante,
                'pasada': True
            })
        else:
            citas_programadas.append({
                'cita': cita,
                'paciente_nombre': paciente.nombre,
                'paciente_edad': edad,
                'paciente_cedula': paciente.cedula,
                'tiempo_restante': tiempo_restante,
                'pasada': False
            })

    return render(request, 'citas_profesional.html', {
        'citas_programadas': citas_programadas
    })

def marcar_asistencia(request, cita_id, estado):
    # Marcar la cita como asistida o inasistida
    cita = Cita.objects.get(id=cita_id)

    # Verificar si ya ha pasado la hora de la cita
    fecha_cita = timezone.make_aware(datetime.combine(cita.dia, cita.hora))
    if timezone.now() > fecha_cita:
        if estado == 'asistida':
            # Actualizar el estado de la cita y la información del paciente
            cita.asistida = True
            cita.save()
            paciente = Usuariosparticipantes.objects.get(cedula=cita.cedula)
            paciente.citas_asistidas += 1
            paciente.save()
        elif estado == 'inasistida':
            # Actualizar el estado de la cita y la información del paciente
            cita.asistida = False
            cita.save()
            paciente = Usuariosparticipantes.objects.get(cedula=cita.cedula)
            paciente.citas_inasistidas += 1
            paciente.save()

        return redirect('citas_profesional')  # Redirigir de nuevo a la vista de citas del profesional
    else:
        # Si la cita aún no ha pasado, no se puede marcar
        return redirect('citas_profesional')  # Redirigir sin cambios


from .forms import CedulaSearchForm
from django.http import JsonResponse

def eliminar_usuario(request):
    if request.method == 'POST':
        # Aquí va tu lógica real para eliminar un usuario
        return JsonResponse({'mensaje': 'Usuario eliminado correctamente'})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def buscar_usuario(request):
    form = CedulaSearchForm(request.GET or None)
    resultados = None
    
    if request.method == "GET" and form.is_valid():
        cedula = form.cleaned_data['cedula']
        
        # Buscamos en todas las bases de datos usando la cédula
        citas = Cita.objects.filter(cedula=cedula)
        usuario = Usuarios.objects.filter(cedula=cedula).first()  # Solo uno
        participante = Usuariosparticipantes.objects.filter(cedula=cedula).first()  # Solo uno
        historial = historialmed.objects.filter(paciente__cedula=cedula)

        # Preparamos los resultados para pasarlos al template
        resultados = {
            'usuario': usuario,
            'citas': citas,
            'participante': participante,
            'historial': historial,
        }

    return render(request, 'buscar_usuario.html', {'form': form, 'resultados': resultados})

from django.contrib.auth.hashers import make_password  # Asegúrate de importar esto para encriptar la contraseña
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def new_patient(request):
    form_participante = newpatient()
    form_historial = HistorialmedForm()
    
    if request.method == 'POST':
        form_participante = newpatient(request.POST)
        form_historial = HistorialmedForm(request.POST)
        
        if form_participante.is_valid() and form_historial.is_valid():
            cedula = form_participante.cleaned_data['cedula']

            # Verificar si la cédula ya existe
            if Usuarios.objects.filter(cedula=cedula).exists():
                return render(request, 'new_patient.html', {
                    'form_usuario': form_participante,
                    'form_historial': form_historial,
                    'mensaje_error': f"La cédula {cedula} ya está registrada."
                })
            
            try:
                with transaction.atomic():
                    # Crear el objeto Usuario y guardarlo
                    print(f"Creando usuario con cédula: {cedula}")
                    usuario = Usuarios(
                        cedula=cedula,
                        nombre="Nombre por defecto",
                        apellidos="Apellidos por defecto",
                        correo=None,
                        tipoa='Participante',
                        direccion=None,
                        fechanacimiento=None
                    )
                    usuario.save()  # Guardamos el objeto Usuario
                    print(f"Usuario guardado con éxito: {usuario}")

                    # Crear el objeto Usuariosparticipantes
                    participante_data = form_participante.cleaned_data
                    participante = Usuariosparticipantes(
                        cedula=participante_data['cedula'],
                        estciv=participante_data['estciv'],
                        sexo=participante_data['sexo'],
                        citas_asistidas=participante_data.get('citas_asistidas', 0),
                        citas_inasistidas=participante_data.get('citas_inasistidas', 0),
                        contactoE1='N/A',
                        NumConE1=None,
                        usuario=usuario  # Relación con el objeto Usuario ya guardado
                    )
                    participante.save()  # Guardamos el objeto Usuariosparticipantes
                    print(f"Participante guardado con éxito: {participante}")

                    # Crear el historial médico
                    historial_data = form_historial.cleaned_data
                    historial = historialmed(
                        paciente=usuario,  # Asignamos el objeto Usuarios relacionado
                        iniciotrat=historial_data['iniciotrat'],
                        ultimocon=historial_data['ultimocon'],
                        recaidas=historial_data['recaidas'],
                        diagnostico=historial_data['diagnostico'],
                        profesional='N/A'
                    )
                    historial.save()  # Guardamos el historial médico
                    print(f"Historial guardado con éxito: Historial de {cedula} con N/A")

                    # Asignar el consumo si existe
                    consumo_value = historial_data.get('consumo')
                    if consumo_value:
                        # Si consumo_value es un QuerySet, selecciona los objetos correctamente
                        if consumo_value.exists():
                            historial.consumo.set(consumo_value)  # Asignamos los objetos Drogas de forma correcta
                            print(f"Consumo asociado con éxito: {consumo_value.first()}")
                        else:
                            print("No se encontraron registros de consumo.")
                    
                return render(request, 'new_patient.html', {
                    'form_usuario': newpatient(),
                    'form_historial': HistorialmedForm(),
                    'mensaje': "Paciente y historial guardados correctamente"
                })

            except Exception as e:
                print(f"Error al guardar: {e}")
                return render(request, 'new_patient.html', {
                    'form_usuario': form_participante,
                    'form_historial': form_historial,
                    'mensaje_error': f"Ocurrió un error al guardar: {e}"
                })
        else:
            print("Los formularios no son válidos")
            print(f"Errores del formulario de participante: {form_participante.errors}")
            print(f"Errores del formulario de historial: {form_historial.errors}")

    return render(request, 'new_patient.html', {
        'form_usuario': form_participante,
        'form_historial': form_historial
    })

import pandas as pd
import pandas as pd
from django.db import transaction
from django.shortcuts import render, redirect
from .models import Usuarios, Usuariosparticipantes, historialmed
from django.shortcuts import redirect
from django.contrib import messages



@csrf_exempt
def import_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "No se proporcionó un archivo Excel.")
            return redirect('new-patient')

        try:
            # Intentamos leer el archivo Excel
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Error al leer el archivo Excel: {e}")
            return redirect('new-patient')

        # Imprimimos las columnas para saber si estamos leyendo correctamente el archivo
        print("Columnas detectadas en el archivo Excel:", df.columns.tolist())

        # Asegúrate de que las columnas tengan los nombres correctos
        expected_columns = ['Cédula paciente', 'Fecha de nacimiento', 'Estado civil', 'Sexo', 'Citas asistidas', 'Citas inasistidas']
        for column in expected_columns:
            if column not in df.columns:
                print(f"[ERROR] Columna esperada {column} no encontrada en el archivo Excel.")
                messages.error(request, f"Columna esperada '{column}' no encontrada en el archivo Excel.")
                return redirect('new-patient')

        creados = 0
        duplicados = 0
        vacios = 0
        errores = 0

        EST_CIV_MAP = {
            "Soltero": "0", "Casado": "1", "Divorciado": "2",
            "union libre": "3", "viudo": "4"
        }

        SEXO_MAP = {
            "Mujer": "0", "Hombre": "1"
        }

        def safe_int(value):
            try:
                return int(value)
            except (ValueError, TypeError):
                return 0

        for index, fila in df.iterrows():
            # Imprimir los valores de la fila para depuración
            print(f"Procesando fila {index + 2}: {fila.tolist()}")

            cedula_raw = fila.get('Cédula paciente')  # Usamos el nombre correcto de la columna
            if pd.isna(cedula_raw) or str(cedula_raw).strip().lower() == 'nan':
                vacios += 1
                continue

            # Normalizamos la cédula: eliminamos espacios y ceros a la izquierda
            cedula = str(cedula_raw).strip().lstrip('0')

            # Verificamos si la cédula ya está registrada
            if Usuarios.objects.filter(cedula=cedula).exists():
                print(f"[DUPLICADO] Ya existe cédula en DB: {cedula}")
                duplicados += 1
                continue

            try:
                # Intentamos convertir la fecha de nacimiento
                fecha_nac = pd.to_datetime(fila.get('Fecha de nacimiento'), errors='coerce')

                # Verificamos si la fecha de nacimiento es válida
                if pd.isna(fecha_nac):
                    print(f"[ERROR] Fecha de nacimiento no válida para cédula: {cedula}")
                    errores += 1
                    continue

                # Creamos el objeto Usuario
                usuario = Usuarios.objects.create(
                    cedula=cedula,
                    nombre="Nombre por defecto",
                    apellidos="Apellidos por defecto",
                    correo=None,
                    tipoa='Participante',
                    direccion=None,
                    fechanacimiento=fecha_nac.date().isoformat() if pd.notna(fecha_nac) else None
                )

                # Creamos el objeto Usuariosparticipantes
                Usuariosparticipantes.objects.create(
                    usuario=usuario,
                    estciv=EST_CIV_MAP.get(str(fila.get('Estado civil')).strip(), "0"),
                    sexo=SEXO_MAP.get(str(fila.get('Sexo')).strip(), "0"),
                    citas_asistidas=safe_int(fila.get('Citas asistidas')),
                    citas_inasistidas=safe_int(fila.get('Citas inasistidas')),
                    contactoE1='N/A',
                    NumConE1=None,
                    cedula=cedula
                )

                print(f"[CREADO] Participante agregado con cédula: {cedula}")
                creados += 1

            except Exception as e:
                errores += 1
                print(f"[ERROR] Fila {index + 2}: {e}")

        # Mostramos los resultados de la importación
        messages.success(
            request,
            f"Importación completada: {creados} creados, {duplicados} duplicados, {vacios} vacíos, {errores} errores."
        )
        return redirect('new-patient')

    return redirect('new-patient')


@csrf_exempt
def ultima_encuesta(request):
    ultima_num_encuesta = PreguntaEncuesta.objects.latest('numero_encuesta').numero_encuesta
    preguntas = PreguntaEncuesta.objects.filter(numero_encuesta=ultima_num_encuesta).order_by('numero_pregunta')

    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('respuesta_'):
                pregunta_id = key.replace('respuesta_', '')
                respuesta_valor = value.strip()

                try:
                    pregunta = PreguntaEncuesta.objects.get(id=pregunta_id)
                    RespuestaUsuario.objects.create(
                        pregunta=pregunta,
                        texto=respuesta_valor
                        # Agrega más campos si los necesitas, como usuario o timestamp
                    )
                except PreguntaEncuesta.DoesNotExist:
                    print(f"[ERROR] Pregunta con ID {pregunta_id} no encontrada")

        return redirect('gracias')  # O cualquier nombre de vista que tengas para agradecer

    return render(request, 'encuestausuario.html', {
        'preguntas': preguntas,
        'numero_encuesta': ultima_num_encuesta,
    })

def home(request):
    return render(request,"login.html")
def politica_privacidad(request):
    return render(request, 'politica_privacidad.html')
from .forms import IdeaFormulario
from .forms import OpcionIdeasForm
from .models import Provoca, PruebaAFavor, PruebaEnContra, Utilidad, IdeaRealista

from .forms import OpcionIdeasForm
from .models import Provoca, PruebaAFavor, PruebaEnContra, Utilidad, IdeaRealista

def agregar_opciones(request):
    if request.method == 'POST':
        form = OpcionIdeasForm(request.POST)
        if form.is_valid():
            tipo = form.cleaned_data['tipo_idea']

            if form.cleaned_data['provoca']:
                Provoca.objects.create(tipo=tipo, texto=form.cleaned_data['provoca'])
            if form.cleaned_data['prueba_a_favor']:
                PruebaAFavor.objects.create(tipo=tipo, texto=form.cleaned_data['prueba_a_favor'])
            if form.cleaned_data['prueba_en_contra']:
                PruebaEnContra.objects.create(tipo=tipo, texto=form.cleaned_data['prueba_en_contra'])
            if form.cleaned_data['utilidad']:
                Utilidad.objects.create(tipo=tipo, texto=form.cleaned_data['utilidad'])
            if form.cleaned_data['idea_realista']:
                IdeaRealista.objects.create(tipo=tipo, texto=form.cleaned_data['idea_realista'])

            return redirect('opcion_agregada')  # Puedes redirigir a una página de éxito
    else:
        form = OpcionIdeasForm()

    return render(request, 'agregar_opciones.html', {'form': form})
@csrf_exempt
def idea_view(request):
    if request.method == 'POST':
        form = IdeaFormulario(request.POST)
        if form.is_valid():
            idea = form.cleaned_data['idea']

            # Guardar opciones personalizadas si se escribieron
            if form.cleaned_data['otra_provoca']:
                Provoca.objects.create(idea=idea, texto=form.cleaned_data['otra_provoca'])
            if form.cleaned_data['otra_prueba_a_favor']:
                PruebaAFavor.objects.create(idea=idea, texto=form.cleaned_data['otra_prueba_a_favor'])
            if form.cleaned_data['otra_prueba_en_contra']:
                PruebaEnContra.objects.create(idea=idea, texto=form.cleaned_data['otra_prueba_en_contra'])
            if form.cleaned_data['otra_utilidad']:
                Utilidad.objects.create(idea=idea, texto=form.cleaned_data['otra_utilidad'])
            if form.cleaned_data['otra_idea_realista']:
                IdeaRealista.objects.create(idea=idea, texto=form.cleaned_data['otra_idea_realista'])

            return render(request, 'resultado.html', {'idea_realista': form.cleaned_data.get('idea_realista') or form.cleaned_data['otra_idea_realista']})
    else:
        form = IdeaFormulario()

    return render(request, 'idea_formulario.html', {'form': form})

def exportar_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"

    columnas = ['Cédula', 'Estado Civil', 'Sexo',
                'Citas Asistidas', 'Citas Inasistidas', 'Fecha de Nacimiento',
                'Inicio de Tratamiento', 'Última Consulta', 'Recaídas', 'Diagnóstico']
    ws.append(columnas)

    participantes = Usuariosparticipantes.objects.all()
    for participante in participantes:
        historial = historialmed.objects.filter(paciente=participante.usuario).first() if participante.usuario else None
        ws.append([
            participante.cedula,
            participante.estciv,
            participante.sexo,
            participante.citas_asistidas,
            participante.citas_inasistidas,
            participante.usuario.fechanacimiento if participante.usuario else '',
            historial.iniciotrat if historial else '',
            historial.ultimocon if historial else '',
            historial.recaidas if historial else '',
            historial.diagnostico if historial else '',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=participantes.xlsx'
    return response

from django.views import View
class DescargarExcelPageView(View):
    def get(self, request, *args, **kwargs):
        participantes = Usuariosparticipantes.objects.all()
        cedulas = [p.cedula for p in participantes]
        asistidas = [p.citas_asistidas for p in participantes]
        inasistidas = [p.citas_inasistidas for p in participantes]

        context = {
            'cedulas': cedulas,
            'asistidas': asistidas,
            'inasistidas': inasistidas,
        }
        return render(request, 'descargar_excel.html', context)

from datetime import datetime

def descargar_datos(request):
    participantes = Usuariosparticipantes.objects.all()

    estciviles = []
    sexos = []
    asistidas = []
    inasistidas = []
    nacimientos = []
    inicios_trat = []
    ultimas_consultas = []
    recaidas = []
    diagnosticos = []
    
    for p in participantes:
        estciviles.append(p.estciv)
        sexos.append(p.sexo)
        asistidas.append(int(p.citas_asistidas or 0))
        inasistidas.append(int(p.citas_inasistidas or 0))
        
        # Fecha de nacimiento
        if p.usuario and p.usuario.fechanacimiento:
            fecha = p.usuario.fechanacimiento
            if isinstance(fecha, str):
                try:
                    fecha = datetime.strptime(fecha, '%Y-%m-%d')
                except ValueError:
                    fecha = None
            if fecha:
                nacimientos.append(fecha.strftime('%Y'))

        historial = historialmed.objects.filter(paciente=p.usuario).first() if p.usuario else None
        if historial:
            # Inicio de tratamiento
            if historial.iniciotrat:
                fecha = historial.iniciotrat
                if isinstance(fecha, str):
                    try:
                        fecha = datetime.strptime(fecha, '%Y-%m-%d')
                    except ValueError:
                        fecha = None
                if fecha:
                    inicios_trat.append(fecha.strftime('%Y'))

            # Última consulta
            if historial.ultimocon:
                fecha = historial.ultimocon
                if isinstance(fecha, str):
                    try:
                        fecha = datetime.strptime(fecha, '%Y-%m-%d')
                    except ValueError:
                        fecha = None
                if fecha:
                    ultimas_consultas.append(fecha.strftime('%Y'))

            if historial.recaidas is not None:
                recaidas.append(int(historial.recaidas))

            if historial.diagnostico:
                diagnosticos.append(historial.diagnostico)

    context = {
        'estciviles': estciviles,
        'sexos': sexos,
        'asistidas': asistidas,
        'inasistidas': inasistidas,
        'nacimientos': nacimientos,
        'inicios_trat': inicios_trat,
        'ultimas_consultas': ultimas_consultas,
        'recaidas': recaidas,
        'diagnosticos': diagnosticos,
    }

    return render(request, 'descargar_excel.html', context)

from .forms import ExcelUploadForm
from .models import ReferenciaCIE10
from django.contrib import messages

def cargar_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            df = pd.read_excel(archivo)

            for _, row in df.iterrows():
                ReferenciaCIE10.objects.update_or_create(
                    codigo=row['Codigo'],
                    defaults={
                        'tabla': row['Tabla'],
                        'nombre': row['Nombre'],
                        'descripcion': row['Descripcion']
                    }
                )
            messages.success(request, "Datos cargados exitosamente.")
            return redirect('buscar_cie10')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'cargar_excel.html', {'form': form})


from django.db.models import Q

def buscar_cie10(request):
    query = request.GET.get('q', '')
    resultados = []

    if query:
        resultados = ReferenciaCIE10.objects.filter(
            Q(descripcion__icontains=query)
        )

    return render(request, 'buscar_cie10.html', {'resultados': resultados, 'query': query})

def regist(request):
    return render(request,"regis.html")

def menu(request):
    if not request.session.get('usuario_id'):
        return redirect('home')  # Redirigir al login si el usuario no está autenticado
    return render(request,"menu.html")

def menu_profesional(request):
    # if not request.session.get('usuario_id'):
    #     return redirect('home')  # Redirigir al login si el usuario no está autenticado
    return render(request,"menu_profesional.html")

def agregar_educativo(request):
    if request.method == 'POST':
        if 'archivo_excel' in request.FILES:
            archivo = request.FILES['archivo_excel']
            try:
                wb = openpyxl.load_workbook(archivo)
                hoja = wb.active

                for fila in hoja.iter_rows(min_row=2, values_only=True):  # Saltamos encabezado
                    numero, tipsdia, link, mensajeprevent, mensajetexto = fila[:5]

                    Educativo.objects.update_or_create(
                        numero=numero,
                        defaults={
                            'tipsdia': tipsdia,
                            'link': link,
                            'mensajeprevent': mensajeprevent,
                            'mensajetexto': mensajetexto
                        }
                    )

                messages.success(request, "Datos cargados desde Excel exitosamente.")
                return redirect('agregar_educativo')

            except Exception as e:
                messages.error(request, f"Error al procesar el Excel: {e}")

        else:
            form = EducativoForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                messages.success(request, "Registro guardado correctamente.")
                return redirect('agregar_educativo')
    else:
        form = EducativoForm()

    return render(request, 'agregar_educativo.html', {'form': form})

@login_required
def misdatosp(request):  

    # Acceder a los datos del usuario desde la sesión
    usuario_nombre = request.session.get('usuario_nombre')
    usuario_apellidos = request.session.get('usuario_apellidos')
    usuario_correo = request.session.get('usuario_correo')
    usuario_direccion = request.session.get('usuario_direccion')
    usuario_fechanacimiento = request.session.get('usuario_fechanacimiento')

    # Puedes pasar estos datos al template
    return render(request, 'infouser.html', {
        'usuario_nombre': usuario_nombre,
        'usuario_apellidos': usuario_apellidos,
        'usuario_correo': usuario_correo,
        'usuario_direccion': usuario_direccion,
        'usuario_fechanacimiento': usuario_fechanacimiento,
    })

def llamadas(request):
    return render(request,"llamadas.html")

from .models import Educativo
def educativ(request):
    educativo = Educativo.objects.first()  # o algún filtro
    return render(request,"moded.html", {'educativo': educativo})

def preguntaseguimiento(request):
    return render(request,"menu.html")

def regiscuenta(request):
    return render(request,"moded.html")

def citas(request):
    return render(request,"menu.html")

def dispc(request):
    return render(request,"menu.html")

from django.shortcuts import render, redirect
from .models import temaforo, DatoSensor
from django.utils import timezone

def foro_view(request):
    if request.method == 'POST':
        nuevo_tema = request.POST.get('tema')
        if nuevo_tema:
            temaforo.objects.create(
                usuario=1,  # Aquí deberías usar `request.user.id` si usas autenticación
                tema=nuevo_tema,
                fechapub=timezone.now(),
                fechaact=timezone.now()
            )
            return redirect('foro')  # Redirige para evitar reenvío de formulario

    temas = temaforo.objects.order_by('-fechaact')
    return render(request, 'foro.html', {'temas': temas})

from .models import DatoSensor, ConnectionSensor
import json


def sensor_ble_view(request):
    # Obtiene el último dato recibido de la base de datos
    # print(DatoSensor.objects.all())
    # ultimo_dato = DatoSensor.objects.last()
    return render(request, 'ble.html')

@csrf_exempt
def agregar_dato_sensor(request):
    if request.method == "POST":
        try:
            user_id = request.session.get('usuario_id')
            body_unicode = request.body.decode('utf-8')
            print("📥 Raw JSON body:", body_unicode)  # Debug
            body = json.loads(body_unicode)
            print("🧾 Parsed body:", body)  # Debug

            temperatura_corporal = float(body.get('temperatura_corporal'))
            temperatura_ambiental = float(body.get('temperatura_ambiental'))
            bpm = int(body.get('bpm'))
            avg_bpm = int(body.get('avg_bpm'))
            humedad = float(body.get('humedad'))
            estado = body.get('estado', 'Conectando')

            ConnectionSensor.objects.create(
                userId = user_id,
                device=body.get("device"),
                temperatura_corporal=temperatura_corporal,
                temperatura_ambiental=temperatura_ambiental,
                bpm=bpm,
                avg_bpm=avg_bpm,
                humedad=humedad,
                estado=estado
            )

            return JsonResponse({"status": "success", "message": "Dato agregado correctamente."})
        except Exception as e:
            print("❌ Exception:", str(e))  # Debug
            return JsonResponse({"status": "error", "message": str(e)})

    return JsonResponse({"status": "error", "message": "Método no permitido"})


@csrf_exempt  # Si no estás utilizando CSRF o si lo manejas en otra parte
def guardar_dato_ble(request):
    user_id = request.session.get('usuario_id')
    print(user_id)
    latest = DatoSensor.objects.last()
    if latest:
        data = {
            'device': latest.device,
            'temperatura_ambiental': latest.temperatura_ambiental,
            'temperatura_corporal': latest.temperatura_corporal,
            'bpm': latest.bpm,
            'avg_bpm': latest.avg_bpm,
            'humedad': latest.humedad,
            'estado': latest.estado,
            'timestamp': latest.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        }
        #print(data)
        return JsonResponse(data)
    else:
        data = {'error': 'No data found'}
        return JsonResponse(data)